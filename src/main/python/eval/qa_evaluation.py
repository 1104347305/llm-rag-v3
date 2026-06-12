from __future__ import annotations

import csv
import asyncio
import io
import json
import math
import os
import re
import sqlite3
import time
import uuid
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from src.main.python.config import settings


ERROR_TYPES = {
    "answer_incorrect",
    "answer_incomplete",
    "question_misunderstood",
    "irrelevant_answer",
    "unsupported_claim",
    "hallucination",
    "context_not_used",
    "insufficient_context",
    "context_noise",
    "over_refusal",
    "format_error",
    "too_vague",
    "contradiction",
    "needs_human_review",
}

MODE_THRESHOLDS = {
    "gold_qa": {"pass": 0.8, "low_confidence": 0.7, "boundary": (0.7, 0.85)},
    "rag_faithfulness": {"pass": 0.75, "low_confidence": 0.65, "boundary": (0.6, 0.8)},
    "weak_qa_quality": {"pass": 0.7, "low_confidence": 0.6, "boundary": (0.55, 0.75)},
}

FIELD_ALIASES = {
    "id": {"id", "样本id", "编号", "序号"},
    "question": {"question", "query", "问题", "用户问题", "提问"},
    "actual_answer": {"actual_answer", "answer", "model_answer", "模型回答", "实际回答", "回答"},
    "gold_answer": {"gold_answer", "expected_answer", "reference_answer", "标准答案", "参考答案"},
    "contexts": {"contexts", "context", "retrieved_context", "上下文", "检索上下文", "证据"},
    "model_name": {"model_name", "model", "模型"},
    "latency_ms": {"latency_ms", "latency", "耗时"},
    "token_usage": {"token_usage", "tokens", "token"},
    "cost": {"cost", "费用"},
    "category": {"category", "类别", "分类", "场景"},
}


@dataclass(frozen=True)
class ParsedUpload:
    rows: list[dict[str, Any]]
    mapping: dict[str, str]


@dataclass(frozen=True)
class MiniMaxJudgeConfig:
    api_key: str
    base_url: str
    model: str
    temperature: float
    max_tokens: int
    timeout: int
    retry_count: int
    workspace: str


def minimax_config() -> MiniMaxJudgeConfig:
    api_key = (
        os.getenv("DASHSCOPE_API_KEY")
        or os.getenv("MINIMAX_API_KEY")
        or settings.dashscope_api_key
        or ""
    ).strip()
    return MiniMaxJudgeConfig(
        api_key=api_key,
        base_url=os.getenv(
            "DASHSCOPE_MINIMAX_BASE_URL",
            os.getenv("MINIMAX_BASE_URL", settings.dashscope_chat_base_url or "https://dashscope.aliyuncs.com/compatible-mode/v1"),
        ).rstrip("/"),
        model=os.getenv("DASHSCOPE_MINIMAX_MODEL", os.getenv("MINIMAX_MODEL", "MiniMax-M2.5")),
        temperature=float(os.getenv("MINIMAX_TEMPERATURE", "0.1")),
        max_tokens=int(os.getenv("MINIMAX_MAX_TOKENS", "1600")),
        timeout=int(os.getenv("MINIMAX_TIMEOUT", "60")),
        retry_count=int(os.getenv("MINIMAX_RETRY_COUNT", "2")),
        workspace=os.getenv("DASHSCOPE_WORKSPACE", "").strip(),
    )


def evaluation_root() -> Path:
    root = settings.storage_dir / "evaluations"
    root.mkdir(parents=True, exist_ok=True)
    return root


def db_path() -> Path:
    return evaluation_root() / "evaluations.sqlite3"


def connect_db() -> sqlite3.Connection:
    conn = sqlite3.connect(db_path())
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    init_db(conn)
    return conn


def init_db(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS evaluation_jobs (
            id TEXT PRIMARY KEY,
            name TEXT NOT NULL,
            status TEXT NOT NULL,
            file_name TEXT NOT NULL,
            total_count INTEGER NOT NULL DEFAULT 0,
            completed_count INTEGER NOT NULL DEFAULT 0,
            failed_count INTEGER NOT NULL DEFAULT 0,
            created_at REAL NOT NULL,
            started_at REAL,
            finished_at REAL,
            config_json TEXT NOT NULL DEFAULT '{}',
            summary_json TEXT NOT NULL DEFAULT '{}',
            threshold_config_json TEXT NOT NULL DEFAULT '{}',
            judge_agreement_json TEXT NOT NULL DEFAULT '{}'
        );

        CREATE TABLE IF NOT EXISTS evaluation_items (
            id TEXT PRIMARY KEY,
            job_id TEXT NOT NULL,
            row_index INTEGER NOT NULL,
            question TEXT NOT NULL DEFAULT '',
            actual_answer TEXT NOT NULL DEFAULT '',
            gold_answer TEXT NOT NULL DEFAULT '',
            contexts_json TEXT NOT NULL DEFAULT '[]',
            category TEXT NOT NULL DEFAULT '',
            evaluation_mode TEXT NOT NULL,
            status TEXT NOT NULL,
            score REAL,
            confidence REAL,
            is_pass INTEGER,
            needs_human_review INTEGER,
            primary_error_type TEXT,
            error_types_json TEXT NOT NULL DEFAULT '[]',
            result_json TEXT NOT NULL DEFAULT '{}',
            raw_judge_response TEXT NOT NULL DEFAULT '',
            schema_warnings_json TEXT NOT NULL DEFAULT '[]',
            failure_reason TEXT NOT NULL DEFAULT '',
            created_at REAL NOT NULL,
            updated_at REAL NOT NULL,
            FOREIGN KEY(job_id) REFERENCES evaluation_jobs(id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS evaluation_item_attempts (
            id TEXT PRIMARY KEY,
            item_id TEXT NOT NULL,
            attempt_no INTEGER NOT NULL,
            status TEXT NOT NULL,
            request_json TEXT NOT NULL DEFAULT '{}',
            response_text TEXT NOT NULL DEFAULT '',
            parsed_json TEXT NOT NULL DEFAULT '{}',
            schema_errors_json TEXT NOT NULL DEFAULT '[]',
            latency_ms INTEGER NOT NULL DEFAULT 0,
            token_usage_json TEXT NOT NULL DEFAULT '{}',
            cost REAL NOT NULL DEFAULT 0,
            created_at REAL NOT NULL,
            FOREIGN KEY(item_id) REFERENCES evaluation_items(id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS evaluation_reviews (
            id TEXT PRIMARY KEY,
            item_id TEXT NOT NULL,
            reviewer TEXT NOT NULL DEFAULT '',
            reviewed_score REAL,
            reviewed_is_pass INTEGER,
            reviewed_primary_error_type TEXT,
            reviewed_error_types_json TEXT NOT NULL DEFAULT '[]',
            reviewed_result_json TEXT NOT NULL DEFAULT '{}',
            review_comment TEXT NOT NULL DEFAULT '',
            is_gold_answer_wrong INTEGER NOT NULL DEFAULT 0,
            is_question_unanswerable INTEGER NOT NULL DEFAULT 0,
            is_context_insufficient INTEGER NOT NULL DEFAULT 0,
            created_at REAL NOT NULL,
            updated_at REAL NOT NULL,
            FOREIGN KEY(item_id) REFERENCES evaluation_items(id) ON DELETE CASCADE
        );
        """
    )
    conn.commit()


def parse_upload(file_name: str, content: bytes) -> ParsedUpload:
    suffix = Path(file_name).suffix.lower()
    if suffix == ".csv":
        rows = _parse_csv(content)
    elif suffix == ".jsonl":
        rows = _parse_jsonl(content)
    elif suffix == ".json":
        rows = _parse_json(content)
    elif suffix == ".xlsx":
        rows = _parse_xlsx(content)
    else:
        raise ValueError("只支持 CSV、XLSX、JSON、JSONL 文件")

    if not rows:
        raise ValueError("文件中没有可读取的数据行")

    mapping = infer_field_mapping(rows[0].keys())
    if "question" not in mapping or "actual_answer" not in mapping:
        raise ValueError("无法识别 question 和 actual_answer 字段，请使用推荐字段名或中文别名")
    return ParsedUpload(rows=rows, mapping=mapping)


def _parse_csv(content: bytes) -> list[dict[str, Any]]:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    return [dict(row) for row in reader]


def _parse_jsonl(content: bytes) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for line in content.decode("utf-8-sig").splitlines():
        if not line.strip():
            continue
        value = json.loads(line)
        if not isinstance(value, dict):
            raise ValueError("JSONL 每一行必须是对象")
        rows.append(value)
    return rows


def _parse_json(content: bytes) -> list[dict[str, Any]]:
    payload = json.loads(content.decode("utf-8-sig"))
    if isinstance(payload, dict) and isinstance(payload.get("data"), list):
        payload = payload["data"]
    if not isinstance(payload, list):
        raise ValueError("JSON 文件必须是对象数组，或包含 data 数组")
    if not all(isinstance(row, dict) for row in payload):
        raise ValueError("JSON 数组中的每一项必须是对象")
    return list(payload)


def _parse_xlsx(content: bytes) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise ValueError("读取 XLSX 需要安装 openpyxl") from exc

    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    parsed: list[dict[str, Any]] = []
    for values in rows[1:]:
        row = {headers[index]: value for index, value in enumerate(values) if index < len(headers) and headers[index]}
        if any(_stringify(value).strip() for value in row.values()):
            parsed.append(row)
    return parsed


def infer_field_mapping(columns: Any) -> dict[str, str]:
    mapping: dict[str, str] = {}
    normalized = {_normalize_column(str(column)): str(column) for column in columns}
    for target, aliases in FIELD_ALIASES.items():
        for alias in aliases:
            found = normalized.get(_normalize_column(alias))
            if found is not None:
                mapping[target] = found
                break
    return mapping


def create_evaluation_job(file_name: str, content: bytes, config: dict[str, Any] | None = None) -> str:
    parsed = parse_upload(file_name, content)
    job_id = uuid.uuid4().hex
    now = time.time()
    config_json = dict(config or {})
    config_json["field_mapping"] = parsed.mapping

    with connect_db() as conn:
        conn.execute(
            """
            INSERT INTO evaluation_jobs (
                id, name, status, file_name, total_count, created_at,
                config_json, threshold_config_json
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                job_id,
                Path(file_name).stem or "QA 评估任务",
                "pending",
                file_name,
                len(parsed.rows),
                now,
                json.dumps(config_json, ensure_ascii=False),
                json.dumps(MODE_THRESHOLDS, ensure_ascii=False),
            ),
        )
        for index, row in enumerate(parsed.rows, start=1):
            item = _normalize_row(row, parsed.mapping)
            mode = detect_evaluation_mode(item)
            conn.execute(
                """
                INSERT INTO evaluation_items (
                    id, job_id, row_index, question, actual_answer, gold_answer,
                    contexts_json, category, evaluation_mode, status, created_at, updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    uuid.uuid4().hex,
                    job_id,
                    index,
                    item["question"],
                    item["actual_answer"],
                    item["gold_answer"],
                    json.dumps(item["contexts"], ensure_ascii=False),
                    item["category"],
                    mode,
                    "pending",
                    now,
                    now,
                ),
            )
        conn.commit()
    return job_id


def run_evaluation_job(job_id: str) -> None:
    with connect_db() as conn:
        job = conn.execute("SELECT * FROM evaluation_jobs WHERE id = ?", (job_id,)).fetchone()
        if job is None:
            raise ValueError(f"job not found: {job_id}")

        started_at = time.time()
        conn.execute("UPDATE evaluation_jobs SET status = ?, started_at = ? WHERE id = ?", ("evaluating", started_at, job_id))
        conn.execute(
            """
            UPDATE evaluation_items
            SET status = ?, updated_at = ?
            WHERE job_id = ? AND status = ?
            """,
            ("pending", started_at, job_id, "evaluating"),
        )
        conn.commit()

        job_config = _json_loads(job["config_json"], {})
        items = conn.execute(
            """
            SELECT * FROM evaluation_items
            WHERE job_id = ? AND status IN ('pending', 'prechecked', 'failed', 'evaluating')
            ORDER BY row_index
            """,
            (job_id,),
        ).fetchall()

        for item in items:
            current_status = conn.execute("SELECT status FROM evaluation_jobs WHERE id = ?", (job_id,)).fetchone()
            if current_status is None:
                raise ValueError(f"job not found: {job_id}")
            if current_status["status"] == "paused":
                summary = build_summary(conn, job_id)
                conn.execute(
                    """
                    UPDATE evaluation_jobs
                    SET completed_count = ?, failed_count = ?, summary_json = ?
                    WHERE id = ?
                    """,
                    (
                        summary["completed_count"],
                        summary["failed_count"],
                        json.dumps(summary, ensure_ascii=False),
                        job_id,
                    ),
                )
                conn.commit()
                return
            _evaluate_item(conn, item, job_config)

        summary = build_summary(conn, job_id)
        remaining_count = conn.execute(
            """
            SELECT COUNT(*) AS count
            FROM evaluation_items
            WHERE job_id = ? AND status NOT IN ('succeeded', 'needs_review', 'reviewed', 'failed', 'skipped')
            """,
            (job_id,),
        ).fetchone()["count"]
        finished_at = time.time()
        final_status = "completed" if remaining_count == 0 else "paused"
        conn.execute(
            """
            UPDATE evaluation_jobs
            SET status = ?, completed_count = ?, failed_count = ?, finished_at = ?, summary_json = ?
            WHERE id = ?
            """,
            (
                final_status,
                summary["completed_count"],
                summary["failed_count"],
                finished_at,
                json.dumps(summary, ensure_ascii=False),
                job_id,
            ),
        )
        conn.commit()


def _evaluate_item(conn: sqlite3.Connection, item: sqlite3.Row, job_config: dict[str, Any] | None = None) -> None:
    started = time.perf_counter()
    item_id = item["id"]
    now = time.time()
    conn.execute("UPDATE evaluation_items SET status = ?, updated_at = ? WHERE id = ?", ("evaluating", now, item_id))
    conn.commit()

    request = {
        "evaluation_mode": item["evaluation_mode"],
        "question": item["question"],
        "actual_answer": item["actual_answer"],
        "gold_answer": item["gold_answer"],
        "contexts": _json_loads(item["contexts_json"], []),
    }

    try:
        result, raw_response, errors, latency_ms = _judge_item_with_attempts(conn, item_id, request, job_config or {})
        conn.execute(
            """
            UPDATE evaluation_items
            SET status = ?, score = ?, confidence = ?, is_pass = ?, needs_human_review = ?,
                primary_error_type = ?, error_types_json = ?, result_json = ?,
                raw_judge_response = ?, schema_warnings_json = ?, failure_reason = ?, updated_at = ?
            WHERE id = ?
            """,
            (
                "needs_review" if result.get("needs_human_review") else "succeeded",
                result["score"],
                result["confidence"],
                int(bool(result["is_pass"])),
                int(bool(result["needs_human_review"])),
                result["primary_error_type"],
                json.dumps(result["error_types"], ensure_ascii=False),
                json.dumps(result, ensure_ascii=False),
                raw_response,
                json.dumps(errors, ensure_ascii=False),
                "",
                time.time(),
                item_id,
            ),
        )
    except Exception as exc:
        latency_ms = round((time.perf_counter() - started) * 1000)
        conn.execute(
            """
            INSERT INTO evaluation_item_attempts (
                id, item_id, attempt_no, status, request_json, schema_errors_json, latency_ms, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                uuid.uuid4().hex,
                item_id,
                1,
                "failed",
                json.dumps(request, ensure_ascii=False),
                json.dumps([str(exc)], ensure_ascii=False),
                latency_ms,
                time.time(),
            ),
        )
        conn.execute(
            "UPDATE evaluation_items SET status = ?, failure_reason = ?, updated_at = ? WHERE id = ?",
            ("failed", str(exc), time.time(), item_id),
        )
    conn.commit()


def pause_evaluation_job(job_id: str) -> bool:
    with connect_db() as conn:
        job = conn.execute("SELECT id, status FROM evaluation_jobs WHERE id = ?", (job_id,)).fetchone()
        if job is None:
            return False
        if job["status"] in {"completed", "failed", "cancelled"}:
            return True
        summary = build_summary(conn, job_id)
        conn.execute(
            """
            UPDATE evaluation_jobs
            SET status = ?, completed_count = ?, failed_count = ?, summary_json = ?
            WHERE id = ?
            """,
            (
                "paused",
                summary["completed_count"],
                summary["failed_count"],
                json.dumps(summary, ensure_ascii=False),
                job_id,
            ),
        )
        conn.commit()
        return True


def _judge_item_with_attempts(
    conn: sqlite3.Connection,
    item_id: str,
    request: dict[str, Any],
    job_config: dict[str, Any],
) -> tuple[dict[str, Any], str, list[str], int]:
    question = _stringify(request.get("question")).strip()
    actual = _stringify(request.get("actual_answer")).strip()
    gold = _stringify(request.get("gold_answer")).strip()
    contexts = request.get("contexts") if isinstance(request.get("contexts"), list) else []
    mode = request["evaluation_mode"]

    precheck = _precheck(question, actual, gold, contexts, mode)
    if precheck is not None:
        result = _apply_thresholds(precheck)
        result["judge_provider"] = "rule_precheck"
        errors = validate_judge_result(result)
        latency_ms = 0
        conn.execute(
            """
            INSERT INTO evaluation_item_attempts (
                id, item_id, attempt_no, status, request_json, response_text,
                parsed_json, schema_errors_json, latency_ms, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                uuid.uuid4().hex,
                item_id,
                0,
                "prechecked",
                json.dumps(request, ensure_ascii=False),
                json.dumps(result, ensure_ascii=False),
                json.dumps(result, ensure_ascii=False),
                json.dumps(errors, ensure_ascii=False),
                latency_ms,
                time.time(),
            ),
        )
        return result, json.dumps(result, ensure_ascii=False), errors, latency_ms

    config = minimax_config()
    if not config.api_key:
        return _local_fallback_result(conn, item_id, request, "缺少 DASHSCOPE_API_KEY，无法调用阿里百炼 MiniMax 单题裁判")

    last_error = "MiniMax judge failed"
    attempts = max(1, config.retry_count + 1)
    for attempt_no in range(1, attempts + 1):
        attempt_started = time.perf_counter()
        prompt = build_minimax_judge_prompt(request, previous_error=last_error if attempt_no > 1 else "")
        raw_response = ""
        parsed: dict[str, Any] = {}
        errors: list[str] = []
        status = "failed"
        try:
            raw_response = asyncio.run(call_minimax_chat(prompt, config))
            parsed = parse_judge_json(raw_response)
            if parsed.get("evaluation_mode") != mode:
                raise ValueError(f"evaluation_mode mismatch: expected {mode}, got {parsed.get('evaluation_mode')}")
            errors = validate_judge_result(parsed)
            if _has_core_schema_error(errors):
                raise ValueError("; ".join(errors))
            parsed = _apply_thresholds(parsed)
            errors = validate_judge_result(parsed)
            status = "succeeded"
            latency_ms = round((time.perf_counter() - attempt_started) * 1000)
            conn.execute(
                """
                INSERT INTO evaluation_item_attempts (
                    id, item_id, attempt_no, status, request_json, response_text,
                    parsed_json, schema_errors_json, latency_ms, created_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    uuid.uuid4().hex,
                    item_id,
                    attempt_no,
                    status,
                    json.dumps({"request": request, "prompt": prompt, "model": config.model}, ensure_ascii=False),
                    raw_response,
                    json.dumps(parsed, ensure_ascii=False),
                    json.dumps(errors, ensure_ascii=False),
                    latency_ms,
                    time.time(),
                ),
            )
            return parsed, raw_response, errors, latency_ms
        except Exception as exc:
            last_error = str(exc)
            latency_ms = round((time.perf_counter() - attempt_started) * 1000)
            if not errors:
                errors = [last_error]
            conn.execute(
                """
                INSERT INTO evaluation_item_attempts (
                    id, item_id, attempt_no, status, request_json, response_text,
                    parsed_json, schema_errors_json, latency_ms, created_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    uuid.uuid4().hex,
                    item_id,
                    attempt_no,
                    status,
                    json.dumps({"request": request, "prompt": prompt, "model": config.model}, ensure_ascii=False),
                    raw_response,
                    json.dumps(parsed, ensure_ascii=False),
                    json.dumps(errors, ensure_ascii=False),
                    latency_ms,
                    time.time(),
                ),
            )
            conn.commit()

    if _allow_local_fallback(job_config):
        return _local_fallback_result(conn, item_id, request, f"MiniMax 单题裁判失败，已重试 {attempts} 次：{last_error}")
    raise RuntimeError(f"MiniMax 单题裁判失败，已重试 {attempts} 次：{last_error}")


def judge_item(payload: dict[str, Any]) -> dict[str, Any]:
    question = _stringify(payload.get("question")).strip()
    actual = _stringify(payload.get("actual_answer")).strip()
    gold = _stringify(payload.get("gold_answer")).strip()
    contexts = payload.get("contexts") if isinstance(payload.get("contexts"), list) else []
    mode = payload["evaluation_mode"]

    precheck = _precheck(question, actual, gold, contexts, mode)
    if precheck is not None:
        return precheck
    if mode == "gold_qa":
        return _judge_gold(question, actual, gold)
    if mode == "rag_faithfulness":
        return _judge_rag(question, actual, contexts)
    return _judge_weak(question, actual)


def _local_fallback_result(
    conn: sqlite3.Connection,
    item_id: str,
    request: dict[str, Any],
    fallback_reason: str,
) -> tuple[dict[str, Any], str, list[str], int]:
    started = time.perf_counter()
    result = judge_item(request)
    result = _apply_thresholds(result)
    result["judge_provider"] = "local_fallback"
    result["fallback_reason"] = fallback_reason
    result["needs_human_review"] = True
    if "needs_human_review" not in result["error_types"]:
        result["error_types"].append("needs_human_review")
    result["reason"] = f"{result.get('reason', '')} 注：MiniMax 调用失败，已使用本地启发式降级评估。"
    errors = validate_judge_result(result)
    latency_ms = round((time.perf_counter() - started) * 1000)
    raw_response = json.dumps(result, ensure_ascii=False)
    conn.execute(
        """
        INSERT INTO evaluation_item_attempts (
            id, item_id, attempt_no, status, request_json, response_text,
            parsed_json, schema_errors_json, latency_ms, created_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            uuid.uuid4().hex,
            item_id,
            999,
            "local_fallback",
            json.dumps({"request": request, "fallback_reason": fallback_reason}, ensure_ascii=False),
            raw_response,
            raw_response,
            json.dumps(errors, ensure_ascii=False),
            latency_ms,
            time.time(),
        ),
    )
    return result, raw_response, errors, latency_ms


def _allow_local_fallback(job_config: dict[str, Any]) -> bool:
    value = job_config.get("allow_local_fallback")
    if value is None:
        value = os.getenv("EVAL_ALLOW_LOCAL_FALLBACK", "true")
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() not in {"0", "false", "no", "off", ""}


def build_minimax_judge_prompt(payload: dict[str, Any], previous_error: str = "") -> str:
    mode = payload["evaluation_mode"]
    mode_instruction = {
        "gold_qa": (
            "你正在执行 Gold QA 评估。必须比较模型回答与标准答案，判断正确性、完整性、关键点覆盖、"
            "错误信息和表达质量。只有该模式可以输出 is_correct。"
        ),
        "rag_faithfulness": (
            "你正在执行 RAG 忠实度评估。只能依据给定 contexts 判断回答是否忠实、是否有无依据断言、"
            "是否利用上下文。不要使用外部知识。该模式输出 is_faithful，不输出 is_correct。"
        ),
        "weak_qa_quality": (
            "你正在执行弱监督 QA 质量评估。没有标准答案或上下文时，只能估计相关性、可用性、逻辑清晰度、"
            "风险控制和表达质量。该模式不能输出准确率结论。"
        ),
    }[mode]
    schema = _mode_schema_instruction(mode)
    retry_note = f"\n上一次输出存在问题：{previous_error}\n请修复后重新输出严格 JSON。" if previous_error else ""
    return f"""
你是严格、保守、可审计的 QA 效果评估裁判。{mode_instruction}

评估规则：
- 只根据输入字段判断，除非题目本身需要常识性语言理解，否则不要补充外部事实。
- 所有分数必须是 0 到 1 之间的小数。
- confidence 表示你对本次判断的把握，不是回答质量。
- error_types 必须是非空数组，且包含 primary_error_type。
- primary_error_type 必须从以下枚举选择：{", ".join(sorted(ERROR_TYPES))}
- reason 必须具体说明判断依据。
- 如果样本证据不足、边界分数、理由不充分或存在高风险，needs_human_review 必须为 true。
- 只能输出一个 JSON 对象，不要输出 Markdown，不要输出解释性前后缀。

输出 JSON 字段要求：
{schema}

待评估样本：
{json.dumps(payload, ensure_ascii=False, indent=2)}
{retry_note}
""".strip()


async def call_minimax_chat(prompt: str, config: MiniMaxJudgeConfig) -> str:
    from openai import AsyncOpenAI

    client = AsyncOpenAI(
        api_key=config.api_key,
        base_url=config.base_url,
        timeout=config.timeout,
        max_retries=config.retry_count,
        default_headers=(
            {"X-DashScope-WorkSpace": config.workspace}
            if config.workspace else None
        ),
    )
    try:
        completion = await client.chat.completions.create(
            model=config.model,
            messages=[
                {"role": "system", "content": "你是 QA 评估裁判，只输出严格 JSON。"},
                {"role": "user", "content": prompt},
            ],
            temperature=config.temperature,
            max_tokens=config.max_tokens,
        )
        content = completion.choices[0].message.content
        if not isinstance(content, str) or not content:
            raise RuntimeError("MiniMax 响应中没有文本内容")
        return content
    finally:
        await client.close()


def parse_judge_json(raw_response: str) -> dict[str, Any]:
    try:
        payload = json.loads(raw_response)
    except json.JSONDecodeError:
        payload = json.loads(_extract_json_object(raw_response))
    if not isinstance(payload, dict):
        raise ValueError("MiniMax 裁判输出不是 JSON 对象")
    return payload


def _extract_json_object(text: str) -> str:
    start = text.find("{")
    end = text.rfind("}")
    if start < 0 or end <= start:
        raise ValueError("无法从 MiniMax 输出中提取 JSON 对象")
    return text[start : end + 1]


def _mode_schema_instruction(mode: str) -> str:
    shared = """
{
  "evaluation_mode": "...",
  "score": 0.0,
  "confidence": 0.0,
  "is_pass": true,
  "needs_human_review": false,
  "primary_error_type": "...",
  "error_types": ["..."],
  "reason": "...",
  "suggested_fix": "..."
}
""".strip()
    extra = {
        "gold_qa": """
额外必须包含：
{
  "is_correct": true,
  "correctness": 0.0,
  "completeness": 0.0,
  "key_point_coverage": 0.0,
  "no_contradiction": 0.0,
  "clarity": 0.0,
  "missing_points": [],
  "wrong_points": []
}
""".strip(),
        "rag_faithfulness": """
额外必须包含：
{
  "is_faithful": true,
  "faithfulness": 0.0,
  "relevance": 0.0,
  "evidence_support": 0.0,
  "hallucination_risk": 0.0,
  "context_usage": 0.0,
  "unsupported_claims": [],
  "used_evidence": [],
  "context_noise": []
}
""".strip(),
        "weak_qa_quality": """
额外必须包含：
{
  "is_usable": true,
  "relevance": 0.0,
  "usefulness": 0.0,
  "coherence": 0.0,
  "risk_control": 0.0,
  "clarity": 0.0,
  "risk_level": "low | medium | high"
}
""".strip(),
    }[mode]
    return f"{shared}\n{extra}"


def _precheck(question: str, actual: str, gold: str, contexts: list[str], mode: str) -> dict[str, Any] | None:
    if not question:
        return _base_result(mode, 0, 0.95, False, True, "format_error", ["format_error"], "问题为空，无法评估。")
    if not actual:
        return _base_result(mode, 0, 0.95, False, True, "answer_incomplete", ["answer_incomplete"], "回答为空。")
    if len(actual) < 4:
        return _base_result(mode, 0.15, 0.85, False, True, "too_vague", ["too_vague"], "回答过短，缺少可评估信息。")
    if _is_garbled(actual):
        return _base_result(mode, 0.1, 0.9, False, True, "format_error", ["format_error"], "回答疑似乱码或重复字符。")
    if mode == "gold_qa" and not gold:
        return _base_result(mode, 0, 0.9, False, True, "format_error", ["format_error"], "Gold QA 模式缺少标准答案。")
    if mode == "rag_faithfulness" and not _join_contexts(contexts).strip():
        return _base_result(mode, 0.35, 0.75, False, True, "insufficient_context", ["insufficient_context"], "RAG 模式缺少可用上下文。")
    return None


def _judge_gold(question: str, actual: str, gold: str) -> dict[str, Any]:
    coverage = _coverage(gold, actual)
    relevance = _coverage(question, actual)
    contradiction = _negation_mismatch(actual, gold)
    completeness = min(1.0, coverage * 1.12)
    correctness = max(0.0, coverage - (0.25 if contradiction else 0))
    no_contradiction = 0.2 if contradiction else 1.0
    clarity = _clarity(actual)
    score = _clamp(correctness * 0.4 + completeness * 0.25 + coverage * 0.2 + no_contradiction * 0.1 + clarity * 0.05)
    confidence = _clamp(0.68 + min(len(gold), 200) / 1000 + min(len(actual), 200) / 1400)
    errors: list[str] = []
    if score < 0.8:
        errors.append("answer_incorrect" if correctness < 0.65 else "answer_incomplete")
    if relevance < 0.15:
        errors.append("irrelevant_answer")
    if contradiction:
        errors.append("contradiction")
    if not errors:
        errors = ["needs_human_review"] if confidence < 0.7 else []
    primary = errors[0] if errors else "needs_human_review"
    missing = _missing_terms(gold, actual)
    reason = f"与标准答案的关键点覆盖约为 {coverage:.0%}，完整性约为 {completeness:.0%}。"
    if contradiction:
        reason += " 回答与标准答案存在疑似肯否矛盾。"
    result = _base_result("gold_qa", score, confidence, score >= MODE_THRESHOLDS["gold_qa"]["pass"], False, primary, errors or [primary], reason)
    result.update(
        {
            "is_correct": score >= MODE_THRESHOLDS["gold_qa"]["pass"],
            "correctness": round(correctness, 4),
            "completeness": round(completeness, 4),
            "key_point_coverage": round(coverage, 4),
            "no_contradiction": round(no_contradiction, 4),
            "clarity": round(clarity, 4),
            "missing_points": missing[:8],
            "wrong_points": ["存在疑似肯否矛盾"] if contradiction else [],
        }
    )
    return _apply_thresholds(result)


def _judge_rag(question: str, actual: str, contexts: list[str]) -> dict[str, Any]:
    context_text = _join_contexts(contexts)
    relevance = max(_coverage(question, actual), _coverage(actual, question))
    evidence_support = _coverage(actual, context_text)
    context_usage = _coverage(context_text[:1000], actual)
    faithfulness = _clamp(evidence_support * 0.75 + context_usage * 0.25)
    unsupported_claims = _unsupported_sentences(actual, context_text)
    hallucination_risk = _clamp(len(unsupported_claims) / max(1, len(_sentences(actual))))
    score = _clamp(faithfulness * 0.35 + relevance * 0.2 + evidence_support * 0.2 + (1 - hallucination_risk) * 0.15 + context_usage * 0.1)
    confidence = _clamp(0.55 + min(len(context_text), 1200) / 4000 + min(len(actual), 240) / 2000)
    errors: list[str] = []
    if evidence_support < 0.18:
        errors.append("unsupported_claim")
    if hallucination_risk >= 0.5:
        errors.append("hallucination")
    if context_usage < 0.08 and evidence_support < 0.3:
        errors.append("context_not_used")
    if relevance < 0.12:
        errors.append("irrelevant_answer")
    if len(context_text) < 20:
        errors.append("insufficient_context")
    primary = errors[0] if errors else "needs_human_review"
    reason = f"回答与上下文的证据匹配约为 {evidence_support:.0%}，问题相关性约为 {relevance:.0%}。"
    if unsupported_claims:
        reason += " 存在未能在上下文中直接匹配的断言。"
    result = _base_result("rag_faithfulness", score, confidence, score >= MODE_THRESHOLDS["rag_faithfulness"]["pass"], False, primary, errors or [primary], reason)
    result.update(
        {
            "is_faithful": score >= MODE_THRESHOLDS["rag_faithfulness"]["pass"] and not {"unsupported_claim", "hallucination"} & set(errors),
            "faithfulness": round(faithfulness, 4),
            "relevance": round(relevance, 4),
            "evidence_support": round(evidence_support, 4),
            "hallucination_risk": round(hallucination_risk, 4),
            "context_usage": round(context_usage, 4),
            "unsupported_claims": unsupported_claims[:5],
            "used_evidence": _matched_context_snippets(actual, context_text),
            "context_noise": [],
        }
    )
    return _apply_thresholds(result)


def _judge_weak(question: str, actual: str) -> dict[str, Any]:
    relevance = max(_coverage(question, actual), _coverage(actual, question))
    usefulness = _clamp(math.log(max(len(actual), 1), 80))
    coherence = _clarity(actual)
    risk_control = 0.35 if _has_absolute_claim(actual) else 0.8
    clarity = _clarity(actual)
    score = _clamp(relevance * 0.3 + usefulness * 0.3 + coherence * 0.2 + risk_control * 0.1 + clarity * 0.1)
    confidence = _clamp(0.42 + min(len(question), 80) / 400 + min(len(actual), 260) / 1600)
    errors: list[str] = []
    if relevance < 0.12:
        errors.append("irrelevant_answer")
    if usefulness < 0.45:
        errors.append("too_vague")
    if risk_control < 0.5:
        errors.append("unsupported_claim")
    primary = errors[0] if errors else "needs_human_review"
    risk_level = "high" if risk_control < 0.5 else "medium" if confidence < 0.6 else "low"
    reason = f"弱监督模式下估计问题相关性约为 {relevance:.0%}，回答可用性约为 {usefulness:.0%}。"
    result = _base_result("weak_qa_quality", score, confidence, score >= MODE_THRESHOLDS["weak_qa_quality"]["pass"], False, primary, errors or [primary], reason)
    result.update(
        {
            "is_usable": score >= MODE_THRESHOLDS["weak_qa_quality"]["pass"],
            "relevance": round(relevance, 4),
            "usefulness": round(usefulness, 4),
            "coherence": round(coherence, 4),
            "risk_control": round(risk_control, 4),
            "clarity": round(clarity, 4),
            "risk_level": risk_level,
        }
    )
    return _apply_thresholds(result)


def validate_judge_result(result: dict[str, Any]) -> list[str]:
    warnings: list[str] = []
    required = [
        "evaluation_mode",
        "score",
        "confidence",
        "is_pass",
        "needs_human_review",
        "primary_error_type",
        "error_types",
        "reason",
    ]
    for field in required:
        if field not in result:
            warnings.append(f"missing_core_field:{field}")
    mode = result.get("evaluation_mode")
    if mode not in MODE_THRESHOLDS:
        warnings.append("invalid_evaluation_mode")
    else:
        mode_required = {
            "gold_qa": ["is_correct", "correctness", "completeness", "key_point_coverage", "no_contradiction", "clarity", "missing_points", "wrong_points"],
            "rag_faithfulness": ["is_faithful", "faithfulness", "relevance", "evidence_support", "hallucination_risk", "context_usage", "unsupported_claims", "used_evidence", "context_noise"],
            "weak_qa_quality": ["is_usable", "relevance", "usefulness", "coherence", "risk_control", "clarity", "risk_level"],
        }[mode]
        for field in mode_required:
            if field not in result:
                warnings.append(f"missing_mode_field:{field}")
    try:
        if not 0 <= float(result.get("score", -1)) <= 1:
            warnings.append("score 越界，已截断到 0-1")
            result["score"] = _clamp(float(result.get("score", 0)))
    except (TypeError, ValueError):
        warnings.append("invalid_score")
    try:
        if not 0 <= float(result.get("confidence", -1)) <= 1:
            warnings.append("confidence 越界，已截断到 0-1")
            result["confidence"] = _clamp(float(result.get("confidence", 0)))
    except (TypeError, ValueError):
        warnings.append("invalid_confidence")
    if result.get("primary_error_type") not in ERROR_TYPES:
        warnings.append("primary_error_type 不在枚举中，已设置为 needs_human_review")
        result["primary_error_type"] = "needs_human_review"
    error_types = result.get("error_types")
    if not isinstance(error_types, list) or not error_types:
        warnings.append("error_types 为空，已补 needs_human_review")
        result["error_types"] = ["needs_human_review"]
    if result["primary_error_type"] not in result["error_types"]:
        result["error_types"].insert(0, result["primary_error_type"])
    if len(_stringify(result.get("reason"))) < 20:
        warnings.append("reason 过短")
        result["needs_human_review"] = True
    return warnings


def _has_core_schema_error(errors: list[str]) -> bool:
    return any(
        error.startswith("missing_core_field:")
        or error.startswith("missing_mode_field:")
        or error in {"invalid_evaluation_mode", "invalid_score", "invalid_confidence"}
        for error in errors
    )


def _apply_thresholds(result: dict[str, Any]) -> dict[str, Any]:
    mode = result["evaluation_mode"]
    threshold = MODE_THRESHOLDS[mode]
    score = _clamp(float(result["score"]))
    confidence = _clamp(float(result["confidence"]))
    low, high = threshold["boundary"]
    serious = score < 0.4 or result["primary_error_type"] in {"hallucination", "answer_incorrect"} or (
        result["primary_error_type"] == "unsupported_claim" and confidence > 0.7
    )
    result["is_pass"] = bool(score >= threshold["pass"] and not serious)
    result["needs_human_review"] = bool(result.get("needs_human_review") or confidence < threshold["low_confidence"] or low <= score <= high or serious)
    if serious and "needs_human_review" not in result["error_types"]:
        result["error_types"].append("needs_human_review")
    result["score"] = round(score, 4)
    result["confidence"] = round(confidence, 4)
    return result


def build_summary(conn: sqlite3.Connection, job_id: str) -> dict[str, Any]:
    items = conn.execute("SELECT * FROM evaluation_items WHERE job_id = ? ORDER BY row_index", (job_id,)).fetchall()
    mode_counts = Counter(item["evaluation_mode"] for item in items)
    completed = [item for item in items if item["status"] in {"succeeded", "needs_review", "reviewed"}]
    failed = [item for item in items if item["status"] == "failed"]
    error_counts: Counter[str] = Counter()
    mode_scores: dict[str, list[float]] = defaultdict(list)
    pass_counts: Counter[str] = Counter()
    needs_review = 0
    severe_items = []
    provider_counts: Counter[str] = Counter()

    for item in completed:
        result = _json_loads(item["result_json"], {})
        provider_counts[result.get("judge_provider", "minimax")] += 1
        mode = item["evaluation_mode"]
        score = float(item["score"] or 0)
        mode_scores[mode].append(score)
        if item["is_pass"]:
            pass_counts[mode] += 1
        if item["needs_human_review"]:
            needs_review += 1
        for error_type in _json_loads(item["error_types_json"], []):
            error_counts[error_type] += 1
        if score < 0.4 or result.get("primary_error_type") in {"hallucination", "answer_incorrect"}:
            severe_items.append(
                {
                    "row_index": item["row_index"],
                    "question": item["question"],
                    "score": score,
                    "primary_error_type": item["primary_error_type"],
                    "reason": result.get("reason", ""),
                }
            )

    total = len(items)
    mode_distribution = {mode: mode_counts.get(mode, 0) for mode in MODE_THRESHOLDS}
    average_scores = {mode: round(sum(scores) / len(scores), 4) if scores else 0 for mode, scores in mode_scores.items()}
    composite = 0.0
    for mode, count in mode_counts.items():
        composite += (sum(mode_scores.get(mode, [])) / max(1, len(mode_scores.get(mode, [])))) * (count / max(1, total))

    return {
        "total_count": total,
        "completed_count": len(completed),
        "failed_count": len(failed),
        "mode_distribution": mode_distribution,
        "average_scores": average_scores,
        "pass_rates": {
            mode: round(pass_counts[mode] / mode_counts[mode], 4) if mode_counts[mode] else 0 for mode in MODE_THRESHOLDS
        },
        "gold_accuracy": round(pass_counts["gold_qa"] / mode_counts["gold_qa"], 4) if mode_counts["gold_qa"] else None,
        "rag_faithful_rate": round(pass_counts["rag_faithfulness"] / mode_counts["rag_faithfulness"], 4) if mode_counts["rag_faithfulness"] else None,
        "weak_usable_rate": round(pass_counts["weak_qa_quality"] / mode_counts["weak_qa_quality"], 4) if mode_counts["weak_qa_quality"] else None,
        "composite_quality_estimate": round(composite, 4),
        "needs_review_count": needs_review,
        "needs_review_rate": round(needs_review / max(1, total), 4),
        "error_distribution": dict(error_counts.most_common()),
        "judge_provider_distribution": dict(provider_counts.most_common()),
        "severe_items": severe_items[:20],
        "warnings": _summary_warnings(total, mode_counts, needs_review),
    }


def get_job(job_id: str) -> dict[str, Any] | None:
    with connect_db() as conn:
        job = conn.execute("SELECT * FROM evaluation_jobs WHERE id = ?", (job_id,)).fetchone()
        if job is None:
            return None
        summary = _json_loads(job["summary_json"], {})
        if not summary:
            summary = build_summary(conn, job_id)
        return _row_to_dict(job) | {"summary": summary}


def list_jobs() -> list[dict[str, Any]]:
    with connect_db() as conn:
        rows = conn.execute("SELECT * FROM evaluation_jobs ORDER BY created_at DESC LIMIT 50").fetchall()
        return [_row_to_dict(row) | {"summary": _json_loads(row["summary_json"], {})} for row in rows]


def get_items(job_id: str, limit: int = 500, offset: int = 0) -> list[dict[str, Any]]:
    with connect_db() as conn:
        rows = conn.execute(
            "SELECT * FROM evaluation_items WHERE job_id = ? ORDER BY row_index LIMIT ? OFFSET ?",
            (job_id, limit, offset),
        ).fetchall()
        items = []
        for row in rows:
            item = _row_to_dict(row)
            item["contexts"] = _json_loads(row["contexts_json"], [])
            item["error_types"] = _json_loads(row["error_types_json"], [])
            item["result"] = _json_loads(row["result_json"], {})
            item.pop("contexts_json", None)
            item.pop("error_types_json", None)
            item.pop("result_json", None)
            items.append(item)
        return items


def build_markdown_report(job_id: str) -> str:
    job = get_job(job_id)
    if job is None:
        raise ValueError("job not found")
    summary = job["summary"]
    lines = [
        f"# QA 评估报告：{job['name']}",
        "",
        "## 1. 评估概览",
        "",
        f"- 总样本数：{summary.get('total_count', 0)}",
        f"- 已完成：{summary.get('completed_count', 0)}",
        f"- 失败：{summary.get('failed_count', 0)}",
        f"- 综合质量估计分：{_percent(summary.get('composite_quality_estimate'))}",
        f"- 人工复核率：{_percent(summary.get('needs_review_rate'))}",
        "",
        "## 2. 样本构成",
        "",
    ]
    for mode, count in summary.get("mode_distribution", {}).items():
        lines.append(f"- {mode}：{count}")
    lines.extend(["", "## 3. 核心指标", ""])
    for key in ("gold_accuracy", "rag_faithful_rate", "weak_usable_rate"):
        value = summary.get(key)
        if value is not None:
            lines.append(f"- {key}：{_percent(value)}")
    lines.extend(["", "## 4. 错误类型分布", ""])
    for error_type, count in summary.get("error_distribution", {}).items():
        lines.append(f"- {error_type}：{count}")
    lines.extend(["", "## 5. 裁判来源", ""])
    for provider, count in summary.get("judge_provider_distribution", {}).items():
        lines.append(f"- {provider}：{count}")
    lines.extend(["", "## 6. 严重错误样本", ""])
    for item in summary.get("severe_items", [])[:20]:
        lines.append(f"- 第 {item['row_index']} 行，score={item['score']}，{item['primary_error_type']}：{item['reason']}")
    lines.extend(["", "## 7. 风险提示", ""])
    for warning in summary.get("warnings", []):
        lines.append(f"- {warning}")
    lines.extend(["", "## 8. 优化建议", ""])
    lines.extend(_optimization_suggestions(summary))
    return "\n".join(lines) + "\n"


def build_csv_export(job_id: str) -> str:
    output = io.StringIO()
    fieldnames = [
        "row_index",
        "evaluation_mode",
        "status",
        "judge_provider",
        "score",
        "confidence",
        "is_pass",
        "needs_human_review",
        "primary_error_type",
        "error_types",
        "question",
        "actual_answer",
        "gold_answer",
        "contexts",
        "reason",
        "suggested_fix",
    ]
    writer = csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()
    for item in get_items(job_id, limit=100000):
        result = item.get("result", {})
        writer.writerow(
            {
                "row_index": item["row_index"],
                "evaluation_mode": item["evaluation_mode"],
                "status": item["status"],
                "judge_provider": result.get("judge_provider", "minimax"),
                "score": item.get("score", ""),
                "confidence": item.get("confidence", ""),
                "is_pass": item.get("is_pass", ""),
                "needs_human_review": item.get("needs_human_review", ""),
                "primary_error_type": item.get("primary_error_type", ""),
                "error_types": json.dumps(item.get("error_types", []), ensure_ascii=False),
                "question": item["question"],
                "actual_answer": item["actual_answer"],
                "gold_answer": item["gold_answer"],
                "contexts": json.dumps(item.get("contexts", []), ensure_ascii=False),
                "reason": result.get("reason", ""),
                "suggested_fix": result.get("suggested_fix", ""),
            }
        )
    return output.getvalue()


def detect_evaluation_mode(row: dict[str, Any]) -> str:
    if row.get("gold_answer"):
        return "gold_qa"
    if row.get("contexts"):
        return "rag_faithfulness"
    return "weak_qa_quality"


def _normalize_row(row: dict[str, Any], mapping: dict[str, str]) -> dict[str, Any]:
    return {
        "question": _stringify(row.get(mapping.get("question", ""), "")),
        "actual_answer": _stringify(row.get(mapping.get("actual_answer", ""), "")),
        "gold_answer": _stringify(row.get(mapping.get("gold_answer", ""), "")),
        "contexts": _parse_contexts(row.get(mapping.get("contexts", ""), "")),
        "category": _stringify(row.get(mapping.get("category", ""), "")),
    }


def _parse_contexts(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, list):
        return [_context_to_text(item) for item in value if _context_to_text(item)]
    if isinstance(value, dict):
        return [_context_to_text(value)]
    text = _stringify(value).strip()
    if not text:
        return []
    try:
        parsed = json.loads(text)
        if isinstance(parsed, list):
            return [_context_to_text(item) for item in parsed if _context_to_text(item)]
        if isinstance(parsed, dict):
            return [_context_to_text(parsed)]
    except json.JSONDecodeError:
        pass
    parts = re.split(r"\n-{3,}\n|\n\n+|[|｜]{3,}", text)
    return [part.strip() for part in parts if part.strip()]


def _context_to_text(value: Any) -> str:
    if isinstance(value, dict):
        source = _stringify(value.get("source", ""))
        content = _stringify(value.get("content", value.get("text", "")))
        return f"{source}: {content}".strip(": ")
    return _stringify(value).strip()


def _base_result(
    mode: str,
    score: float,
    confidence: float,
    is_pass: bool,
    needs_review: bool,
    primary: str,
    errors: list[str],
    reason: str,
) -> dict[str, Any]:
    return {
        "evaluation_mode": mode,
        "score": _clamp(score),
        "confidence": _clamp(confidence),
        "is_pass": is_pass,
        "needs_human_review": needs_review,
        "primary_error_type": primary,
        "error_types": errors,
        "reason": reason,
        "suggested_fix": _suggested_fix(primary),
    }


def _coverage(source: str, target: str) -> float:
    source_terms = _terms(source)
    target_terms = set(_terms(target))
    if not source_terms:
        return 0.0
    return len([term for term in source_terms if term in target_terms]) / len(source_terms)


def _terms(text: str) -> list[str]:
    text = _stringify(text).lower()
    words = re.findall(r"[a-z0-9_]+|[\u4e00-\u9fff]{1,2}", text)
    stop = {"的", "了", "和", "与", "及", "是", "在", "为", "可以", "什么", "如何", "多少"}
    return [word for word in words if word not in stop and len(word.strip()) > 0]


def _sentences(text: str) -> list[str]:
    return [part.strip() for part in re.split(r"[。！？!?；;\n]+", text) if part.strip()]


def _unsupported_sentences(actual: str, context: str) -> list[str]:
    unsupported = []
    for sentence in _sentences(actual):
        if len(sentence) < 8:
            continue
        if _coverage(sentence, context) < 0.18 and _has_assertion(sentence):
            unsupported.append(sentence)
    return unsupported


def _matched_context_snippets(actual: str, context: str) -> list[str]:
    snippets = []
    actual_terms = set(_terms(actual))
    for sentence in _sentences(context):
        terms = set(_terms(sentence))
        if terms and len(terms & actual_terms) / len(terms) > 0.35:
            snippets.append(sentence[:160])
        if len(snippets) >= 5:
            break
    return snippets


def _missing_terms(gold: str, actual: str) -> list[str]:
    actual_terms = set(_terms(actual))
    missing = []
    for term, count in Counter(_terms(gold)).most_common():
        if term not in actual_terms and term not in missing:
            missing.append(term)
    return missing


def _clarity(text: str) -> float:
    if _is_garbled(text):
        return 0.15
    length = len(text)
    if length < 10:
        return 0.35
    if length > 2000:
        return 0.65
    punctuation_bonus = 0.1 if re.search(r"[。！？；,.!?;]", text) else 0
    return _clamp(0.68 + punctuation_bonus + min(length, 300) / 1500)


def _is_garbled(text: str) -> bool:
    compact = re.sub(r"\s+", "", text)
    if not compact:
        return False
    unique_ratio = len(set(compact)) / len(compact)
    return len(compact) > 10 and unique_ratio < 0.18


def _negation_mismatch(actual: str, gold: str) -> bool:
    negations = {"不", "无", "未", "不能", "不可", "不是", "没有"}
    return any(word in actual for word in negations) != any(word in gold for word in negations)


def _has_assertion(text: str) -> bool:
    return bool(re.search(r"一定|必须|所有|全部|绝对|肯定|不会|总是|唯一|最高|最低", text)) or len(text) > 20


def _has_absolute_claim(text: str) -> bool:
    return bool(re.search(r"所有|任何|绝对|一定|完全|永远|唯一|100%|百分之百", text))


def _join_contexts(contexts: list[str]) -> str:
    return "\n".join(_stringify(context) for context in contexts if _stringify(context).strip())


def _summary_warnings(total: int, mode_counts: Counter[str], needs_review: int) -> list[str]:
    warnings = []
    if total and mode_counts["weak_qa_quality"] / total > 0.5:
        warnings.append("当前综合质量估计分主要基于弱监督判断，不能等同于模型准确率。")
    if total and needs_review / total > 0.3:
        warnings.append("低置信度或需人工复核样本占比较高，建议扩大人工复核样本。")
    return warnings


def _optimization_suggestions(summary: dict[str, Any]) -> list[str]:
    errors = summary.get("error_distribution", {})
    suggestions = []
    if errors.get("unsupported_claim", 0) or errors.get("hallucination", 0):
        suggestions.append("- P0：减少无依据断言。建议要求回答逐条引用上下文，缺少依据时明确说明无法确定。")
    if errors.get("answer_incomplete", 0) or errors.get("too_vague", 0):
        suggestions.append("- P1：提升答案完整性。建议在生成 prompt 中要求覆盖关键条件、次数、对象、限制和例外。")
    if errors.get("context_not_used", 0) or errors.get("insufficient_context", 0):
        suggestions.append("- P1：优化检索链路。建议检查 query rewrite、召回 top_k、chunk 粒度和重排序配置。")
    if not suggestions:
        suggestions.append("- P2：当前未发现集中的严重错误，可优先补充 Gold QA 样本提升评估可信度。")
    return suggestions


def _suggested_fix(primary: str) -> str:
    fixes = {
        "answer_incorrect": "补充标准答案对齐样本，检查知识库来源和回答 prompt。",
        "answer_incomplete": "要求回答覆盖关键条件、适用对象、限制和例外。",
        "unsupported_claim": "要求每个关键结论引用上下文，缺少依据时使用不确定表达。",
        "hallucination": "增强忠实度约束，禁止基于上下文之外的信息扩展。",
        "context_not_used": "检查上下文注入格式，并要求模型优先依据检索内容回答。",
        "insufficient_context": "优化检索召回，补充知识库缺失内容。",
        "irrelevant_answer": "检查问题理解、query rewrite 和 prompt 中的回答范围约束。",
        "too_vague": "要求回答给出具体条件、步骤或结论，避免泛泛而谈。",
        "format_error": "检查输入数据格式和字段映射。",
    }
    return fixes.get(primary, "建议进入人工复核，确认样本和裁判判断是否可靠。")


def _percent(value: Any) -> str:
    if value is None:
        return "-"
    return f"{float(value) * 100:.1f}%"


def _json_loads(text: str, default: Any) -> Any:
    try:
        return json.loads(text)
    except (TypeError, json.JSONDecodeError):
        return default


def _row_to_dict(row: sqlite3.Row) -> dict[str, Any]:
    return {key: row[key] for key in row.keys()}


def _normalize_column(value: str) -> str:
    return re.sub(r"[\s_\-]+", "", value.strip().lower())


def _stringify(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def _clamp(value: float) -> float:
    return max(0.0, min(1.0, float(value)))
