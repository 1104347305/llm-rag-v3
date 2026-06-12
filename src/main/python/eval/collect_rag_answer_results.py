from __future__ import annotations

import argparse
import csv
import json
import re
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


DEFAULT_STREAM_URL = "http://localhost:8010/rag/chat/stream"
DEFAULT_JSON_URL = "http://localhost:8010/rag/answer"
DEFAULT_PROJECT_ID = "quanyi_wiki"
CITATION_PATTERN = re.compile(r"(?<!\w)\[(\d+)\]")
HIDDEN_CITED_PATTERN = re.compile(r"<!--\s*cited:\s*([0-9,\s]+)\s*-->", re.IGNORECASE)


def main() -> None:
    args = parse_args()
    queries = load_queries(args.input, query_column=args.query_column)
    if not queries:
        raise SystemExit(f"No queries found in {args.input}")

    timestamp = time.strftime("%Y%m%d_%H%M%S")
    output_path = add_timestamp(Path(args.output), timestamp)
    if args.excel_output:
        excel_output_path = add_timestamp(Path(args.excel_output), timestamp)
    else:
        excel_output_path = output_path.with_suffix(".xlsx")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    excel_output_path.parent.mkdir(parents=True, exist_ok=True)

    url = args.url or (DEFAULT_STREAM_URL if args.mode == "stream" else DEFAULT_JSON_URL)
    payload_overrides = parse_payload_overrides(args.payload_json)
    started = time.perf_counter()
    rows_by_index: dict[int, dict[str, Any]] = {}

    with output_path.open("w", encoding="utf-8") as jsonl_file:
        with ThreadPoolExecutor(max_workers=args.workers) as executor:
            futures = []
            for index, query in enumerate(queries, start=1):
                futures.append(
                    executor.submit(
                        collect_one,
                        index=index,
                        url=url,
                        project_id=args.project_id,
                        query=query,
                        timeout=args.timeout,
                        payload_overrides=payload_overrides,
                        mode=args.mode,
                    )
                )
                if args.submit_sleep > 0 and index < len(queries):
                    time.sleep(args.submit_sleep)

            for done_count, future in enumerate(as_completed(futures), start=1):
                row = future.result()
                rows_by_index[int(row["input_index"])] = row
                jsonl_file.write(json.dumps(row, ensure_ascii=False) + "\n")
                jsonl_file.flush()
                print_progress(done_count, len(queries), row)

    rows = [rows_by_index[index] for index in sorted(rows_by_index)]
    write_excel(excel_output_path, rows)
    total_elapsed_ms = round((time.perf_counter() - started) * 1000, 2)
    print(f"Saved JSONL: {output_path}")
    print(f"Saved Excel: {excel_output_path}")
    print(f"Completed {len(rows)}/{len(queries)} queries in {total_elapsed_ms}ms")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Concurrently call /rag/answer/stream and save query, answer, TTFT, total time, citations, and raw metadata."
    )
    parser.add_argument("--input", default='/Users/mickey/project/PA-ALG/llm-wiki-rag/docs/权益评估集.xlsx', help="Input file: .txt, .csv, .jsonl, .json, or .xlsx.")
    parser.add_argument(
        "--output",
        default="outputs/rag_answer_results.jsonl",
        help="JSONL output path. A timestamp is appended before the extension.",
    )
    parser.add_argument(
        "--excel-output",
        default="/Users/mickey/project/PA-ALG/llm-wiki-rag/docs/权益评估集_result.xlsx",
        help="Excel output path. A timestamp is appended before the extension.",
    )
    parser.add_argument("--url", default="http://localhost:8010/rag/chat/stream", help="Endpoint URL. Defaults to stream or json URL according to --mode.")
    parser.add_argument("--mode", choices=["stream", "json"], default="stream", help="Use SSE streaming endpoint or legacy JSON endpoint.")
    parser.add_argument("--workers", type=int, default=4, help="Concurrent request count.")
    parser.add_argument("--project-id", default=DEFAULT_PROJECT_ID, help="RAG project_id for every request.")
    parser.add_argument("--query-column", default="question", help="Query column/key for CSV, JSON, JSONL, or XLSX inputs.")
    parser.add_argument("--timeout", type=float, default=180.0, help="Socket timeout in seconds for each request.")
    parser.add_argument("--submit-sleep", type=float, default=0.0, help="Sleep seconds between submitting requests.")
    parser.add_argument("--sleep", dest="submit_sleep", type=float, help=argparse.SUPPRESS)
    parser.add_argument(
        "--payload-json",
        default=None,
        help='Extra request fields as JSON, for example: \'{"include_es": false, "top_pages": 8}\'.',
    )
    return parser.parse_args()


def add_timestamp(path: Path, timestamp: str) -> Path:
    suffix = path.suffix
    if suffix:
        return path.with_name(f"{path.stem}_{timestamp}{suffix}")
    return path.with_name(f"{path.name}_{timestamp}")


def load_queries(input_path: str, query_column: str = "query") -> list[str]:
    path = Path(input_path)
    suffix = path.suffix.lower()
    if suffix in {"", ".txt"}:
        return load_txt_queries(path)
    if suffix == ".csv":
        return load_csv_queries(path, query_column)
    if suffix == ".jsonl":
        return load_jsonl_queries(path, query_column)
    if suffix == ".json":
        return load_json_queries(path, query_column)
    if suffix == ".xlsx":
        return load_xlsx_queries(path, query_column)
    raise ValueError(f"Unsupported input format: {path.suffix}")


def load_txt_queries(path: Path) -> list[str]:
    return [line.strip() for line in path.read_text(encoding="utf-8").splitlines() if line.strip()]


def load_csv_queries(path: Path, query_column: str) -> list[str]:
    with path.open("r", encoding="utf-8-sig", newline="") as file:
        reader = csv.DictReader(file)
        if not reader.fieldnames:
            return []
        column = resolve_column(reader.fieldnames, query_column)
        return [str(row.get(column, "")).strip() for row in reader if str(row.get(column, "")).strip()]


def load_jsonl_queries(path: Path, query_column: str) -> list[str]:
    queries: list[str] = []
    with path.open("r", encoding="utf-8") as file:
        for line in file:
            line = line.strip()
            if not line:
                continue
            data = json.loads(line)
            queries.append(extract_query(data, query_column))
    return [query for query in queries if query]


def load_json_queries(path: Path, query_column: str) -> list[str]:
    data = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(data, list):
        return [query for item in data if (query := extract_query(item, query_column))]
    query = extract_query(data, query_column)
    return [query] if query else []


def load_xlsx_queries(path: Path, query_column: str) -> list[str]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("Reading .xlsx requires openpyxl: python3 -m pip install openpyxl") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(rows, [])]
    column = resolve_column(headers, query_column)
    column_index = headers.index(column)
    queries: list[str] = []
    for row in rows:
        if column_index >= len(row):
            continue
        value = row[column_index]
        if value is not None and str(value).strip():
            queries.append(str(value).strip())
    return queries


def resolve_column(fieldnames: list[str], query_column: str) -> str:
    if query_column in fieldnames:
        return query_column
    normalized = {name.strip().lower(): name for name in fieldnames}
    for candidate in [query_column, "query", "question", "问题", "用户问题"]:
        found = normalized.get(candidate.strip().lower())
        if found:
            return found
    raise ValueError(f"Cannot find query column {query_column!r}. Available columns: {fieldnames}")


def extract_query(data: Any, query_column: str) -> str:
    if isinstance(data, str):
        return data.strip()
    if isinstance(data, dict):
        for key in [query_column, "query", "question", "问题", "用户问题"]:
            value = data.get(key)
            if value is not None and str(value).strip():
                return str(value).strip()
    return ""


def collect_one(
    index: int,
    url: str,
    project_id: str,
    query: str,
    timeout: float,
    payload_overrides: dict[str, Any],
    mode: str,
) -> dict[str, Any]:
    if mode == "stream":
        return collect_one_stream(index, url, project_id, query, timeout, payload_overrides)
    return collect_one_json(index, url, project_id, query, timeout, payload_overrides)


def collect_one_stream(
    index: int,
    url: str,
    project_id: str,
    query: str,
    timeout: float,
    payload_overrides: dict[str, Any],
) -> dict[str, Any]:
    payload = {
        "source": project_id,
        "user_text": query,
        "session_id": f"11111111",
        "user_id": "eval_user",
        "trace_id": f"2222222",
        **payload_overrides,
    }
    started = time.perf_counter()
    status_code: int | None = None
    error: str | None = None
    last_event: dict[str, Any] = {}
    first_event_ms: float | None = None
    client_ttft_ms: float | None = None
    response_open_ms: float | None = None
    event_count = 0

    try:
        req = build_json_request(url, payload, accept="text/event-stream")
        with urlopen(req, timeout=timeout) as response:
            status_code = response.status
            response_open_ms = elapsed_ms_since(started)
            for event in iter_sse_events(response):
                now_ms = elapsed_ms_since(started)
                if first_event_ms is None:
                    first_event_ms = now_ms
                event_count += 1

                data = event.get("data", {}) if isinstance(event.get("data"), dict) else {}
                robot_text = str(data.get("robot_text") or "")
                if robot_text and client_ttft_ms is None:
                    client_ttft_ms = now_ms

                last_event = event
    except HTTPError as exc:
        status_code = exc.code
        error = exc.read().decode("utf-8", errors="replace")
    except (URLError, TimeoutError, json.JSONDecodeError) as exc:
        error = str(exc)

    total_elapsed_ms = elapsed_ms_since(started)
    row = build_result_row(
        input_index=index,
        query=query,
        elapsed_ms=total_elapsed_ms,
        status_code=status_code,
        response_event=last_event,
        error=error,
    )
    row.update({
        "mode": "stream",
        "client_response_open_ms": response_open_ms,
        "client_first_event_ms": first_event_ms,
        "client_ttft_ms": client_ttft_ms,
        "event_count": event_count,
    })
    return row


def collect_one_json(
    index: int,
    url: str,
    project_id: str,
    query: str,
    timeout: float,
    payload_overrides: dict[str, Any],
) -> dict[str, Any]:
    payload = {
        "source": project_id,
        "user_text": query,
        "session_id": f"11111111",
        "user_id": "eval_user",
        "trace_id": f"2222222",
        **payload_overrides,
    }
    started = time.perf_counter()
    status_code: int | None = None
    response_data: dict[str, Any] = {}
    error: str | None = None
    response_open_ms: float | None = None

    try:
        req = build_json_request(url, payload, accept="application/json")
        with urlopen(req, timeout=timeout) as response:
            status_code = response.status
            response_open_ms = elapsed_ms_since(started)
            response_data = json.loads(response.read().decode("utf-8"))
    except HTTPError as exc:
        status_code = exc.code
        error = exc.read().decode("utf-8", errors="replace")
    except (URLError, TimeoutError, json.JSONDecodeError) as exc:
        error = str(exc)

    total_elapsed_ms = elapsed_ms_since(started)
    row = build_result_row(
        input_index=index,
        query=query,
        elapsed_ms=total_elapsed_ms,
        status_code=status_code,
        response_event={"data": response_data.get("data", {})},
        error=error,
    )
    row.update({
        "mode": "json",
        "client_response_open_ms": response_open_ms,
        "client_first_event_ms": None,
        "client_ttft_ms": None,
        "event_count": None,
    })
    return row


def build_json_request(url: str, payload: dict[str, Any], accept: str) -> Request:
    body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    return Request(
        url,
        data=body,
        method="POST",
        headers={"Content-Type": "application/json", "Accept": accept},
    )


def iter_sse_events(response: Any):
    data_lines: list[str] = []
    for raw_line in response:
        line = raw_line.decode("utf-8", errors="replace").rstrip("\r\n")
        if not line:
            if data_lines:
                yield parse_sse_data("\n".join(data_lines))
                data_lines = []
            continue
        if line.startswith("data:"):
            data_lines.append(line[5:].lstrip())
    if data_lines:
        yield parse_sse_data("\n".join(data_lines))


def parse_sse_data(data: str) -> dict[str, Any]:
    payload = json.loads(data)
    if not isinstance(payload, dict):
        return {"type": "unknown", "data": payload}
    return payload


def build_result_row(
    input_index: int,
    query: str,
    elapsed_ms: float,
    status_code: int | None,
    response_event: dict[str, Any],
    error: str | None,
) -> dict[str, Any]:
    data = response_event.get("data", {}) if isinstance(response_event.get("data"), dict) else {}
    answer = str(data.get("robot_text") or "")
    extra = data.get("extra_output_params", {}) if isinstance(data.get("extra_output_params"), dict) else {}
    rewritten_query = extra.get("rewritten_query") or ""
    pages = extra.get("pages") if isinstance(extra.get("pages"), list) else []
    citations = extract_citations(pages)
    cited_numbers = extract_cited_numbers(answer)
    cited_pages = [item for item in citations if item.get("number") in cited_numbers]
    return {
        "input_index": input_index,
        "query": query,
        "rewritten_query": rewritten_query,
        "answer": strip_hidden_cited_comment(answer),
        "elapsed_ms": elapsed_ms,
        "ok": error is None and status_code is not None and 200 <= status_code < 300,
        "status_code": status_code,
        "error": error,
        "llm_error": None,
        "session_id": None,
        "agent_engine": "askbob",
        "cited_numbers": cited_numbers,
        "cited_pages": cited_pages,
        "citations": citations,
        "fallback_reasons": [],
        "server_total_ms": extra.get("final_frame_time"),
        "server_retrieval_ms": extra.get("first_frame_time"),
        "server_llm_ms": None,
        "metrics": extra.get("metrics", {}),
        "raw_response": response_event,
    }


def extract_citations(pages: list[Any]) -> list[dict[str, Any]]:
    citations: list[dict[str, Any]] = []
    for page in pages:
        if not isinstance(page, dict):
            continue
        citations.append(
            {
                "number": page.get("number"),
                "title": page.get("title"),
                "path": page.get("path"),
                "page_id": page.get("page_id"),
                "score": page.get("score"),
                "source": page.get("source"),
                "chunk_ids": [
                    chunk.get("chunk_id")
                    for chunk in page.get("chunks", [])
                    if isinstance(chunk, dict) and chunk.get("chunk_id")
                ],
            }
        )
    return citations


def extract_cited_numbers(answer: str) -> list[int]:
    numbers = {int(match.group(1)) for match in CITATION_PATTERN.finditer(answer)}
    for match in HIDDEN_CITED_PATTERN.finditer(answer):
        numbers.update(int(item) for item in re.findall(r"\d+", match.group(1)))
    return sorted(numbers)


def strip_hidden_cited_comment(answer: str) -> str:
    return HIDDEN_CITED_PATTERN.sub("", answer).strip()


def parse_payload_overrides(payload_json: str | None) -> dict[str, Any]:
    if not payload_json:
        return {}
    data = json.loads(payload_json)
    if not isinstance(data, dict):
        raise ValueError("--payload-json must decode to a JSON object")
    return data


def write_excel(path: Path, rows: list[dict[str, Any]]) -> None:
    if not HAS_OPENPYXL:
        raise RuntimeError("Writing .xlsx requires openpyxl: python3 -m pip install openpyxl")

    fieldnames = [
        "input_index",
        "query",
        "rewritten_query",
        "answer",
        "mode",
        "elapsed_ms",
        "client_response_open_ms",
        "client_first_event_ms",
        "client_ttft_ms",
        "server_total_ms",
        "server_retrieval_ms",
        "server_llm_ms",
        "event_count",
        "answer_chars",
        "ok",
        "status_code",
        "error",
        "llm_error",
        "agent_engine",
        "session_id",
        "cited_numbers",
        "cited_pages",
        "citations",
        "fallback_reasons",
        "metrics",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "RAG评估结果"

    # 表头样式
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(vertical="top", wrap_text=True)

    # 写表头
    for col_idx, field in enumerate(fieldnames, start=1):
        cell = ws.cell(row=1, column=col_idx, value=field)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # 写数据
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, field in enumerate(fieldnames, start=1):
            value = row.get(field)
            if value is None:
                cell_value = ""
            elif isinstance(value, (dict, list)):
                cell_value = json.dumps(value, ensure_ascii=False)
            else:
                cell_value = str(value)
            cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
            cell.alignment = cell_alignment

    # 自动列宽
    for col_idx, field in enumerate(fieldnames, start=1):
        max_width = len(field) + 4
        for row_idx in range(2, min(len(rows) + 2, 52)):  # 采样前50行
            cell_value = str(ws.cell(row=row_idx, column=col_idx).value or "")
            max_width = max(max_width, min(len(cell_value), 60))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_width

    # 冻结首行
    ws.freeze_panes = "A2"

    # 添加汇总 sheet
    ws2 = wb.create_sheet("汇总")
    summary_data = {
        "指标": ["总请求数", "成功数", "失败数", "成功率", "平均总耗时(ms)", "平均TTFT(ms)", "平均检索耗时(ms)", "平均LLM耗时(ms)", "平均事件数", "平均答案字数"],
        "值": build_summary(rows, fieldnames),
    }
    for col_idx, header in enumerate(summary_data.keys(), start=1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
    for row_idx, key in enumerate(summary_data["指标"], start=2):
        ws2.cell(row=row_idx, column=1, value=key).font = Font(bold=True)
        ws2.cell(row=row_idx, column=2, value=summary_data["值"][row_idx - 2])
    ws2.column_dimensions["A"].width = 24
    ws2.column_dimensions["B"].width = 20

    wb.save(path)


def build_summary(rows: list[dict[str, Any]], fieldnames: list[str]) -> list[Any]:
    total = len(rows)
    ok_count = sum(1 for r in rows if r.get("ok"))
    fail_count = total - ok_count
    rate = f"{ok_count / total * 100:.1f}%" if total > 0 else "N/A"

    def avg(key: str) -> str:
        vals = [r.get(key) for r in rows if r.get(key) is not None]
        return f"{sum(vals) / len(vals):.1f}" if vals else "N/A"

    return [
        total, ok_count, fail_count, rate,
        avg("elapsed_ms"), avg("client_ttft_ms"),
        avg("server_retrieval_ms"), avg("server_llm_ms"),
        avg("event_count"), avg("answer_chars"),
    ]


def elapsed_ms_since(started: float) -> float:
    return round((time.perf_counter() - started) * 1000, 2)


def print_progress(done_count: int, total: int, row: dict[str, Any]) -> None:
    status = "OK" if row["ok"] else "ERR"
    index = row.get("input_index")
    ttft = row.get("client_ttft_ms") or ""
    retrieval = row.get("server_retrieval_ms") or ""
    llm = row.get("server_llm_ms") or ""

    # 进度条
    bar_width = 30
    filled = int(bar_width * done_count / total)
    bar = "█" * filled + "░" * (bar_width - filled)

    timing_parts = [f"total={row['elapsed_ms']}ms"]
    if ttft:
        timing_parts.append(f"ttft={ttft}ms")
    if retrieval:
        timing_parts.append(f"retrieval={retrieval}ms")
    if llm:
        timing_parts.append(f"llm={llm}ms")

    answer = str(row.get("answer") or "")
    answer_preview = answer[:120].replace("\n", " ") + ("..." if len(answer) > 120 else "")

    print(f"\n{'─' * 60}")
    print(f"[{done_count}/{total}] {bar} {done_count * 100 // total}% | #{index} {status} | {' '.join(timing_parts)}")
    print(f"  问题: {row['query']}")
    print(f"  答案: {answer_preview}")


if __name__ == "__main__":
    main()
