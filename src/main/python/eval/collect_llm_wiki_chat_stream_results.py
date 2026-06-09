from __future__ import annotations

import argparse
import csv
import json
import os
import sys
import time
from pathlib import Path
from typing import Any

try:
    import requests
except ModuleNotFoundError:
    requests = None  # type: ignore[assignment]

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ModuleNotFoundError as exc:
    raise RuntimeError(
        "This script requires openpyxl: python3 -m pip install openpyxl"
    ) from exc


os.environ["PYTHONIOENCODING"] = "utf-8"
for _stream in (sys.stdout, sys.stderr):
    if hasattr(_stream, "reconfigure"):
        _stream.reconfigure(encoding="utf-8", errors="replace")


BASE_URL = os.getenv("LLM_WIKI_BASE_URL", "http://8.148.158.241:8081")
PROJECT_PATH = os.getenv("LLM_WIKI_PROJECT_PATH", "/llm-wiki-7777/wiki-data/rag")

QUESTION_ALIASES = ("question", "query", "问题", "用户问题", "提问")
GOLD_ANSWER_ALIASES = (
    "gold_answer",
    "expected_answer",
    "reference_answer",
    "标准答案",
    "参考答案",
)
CATEGORY_ALIASES = ("category", "type", "分类", "类别", "问题类型")


def chat_stream(
    question: str,
    history: list[dict] | None = None,
    top_k: int = 8,
) -> dict:
    """Call the LLM Wiki SSE API, print tokens, and return the complete result."""
    if requests is None:
        return {
            "answer": "",
            "sources": [],
            "retrieval_query": "",
            "retrieval_ms": 0,
            "error": "缺少 requests 依赖，请执行：python3 -m pip install requests",
        }

    messages = list(history) if history else []
    messages.append({"role": "user", "content": question})

    try:
        response = requests.post(
            f"{BASE_URL}/api/chat/stream",
            json={
                "project_path": PROJECT_PATH,
                "messages": messages,
                "top_k": top_k,
                "stream": True,
            },
            timeout=90,
            stream=True,
        )
        response.raise_for_status()
    except requests.RequestException as exc:
        return {
            "answer": "",
            "sources": [],
            "retrieval_query": "",
            "retrieval_ms": 0,
            "error": str(exc),
        }

    meta: dict[str, Any] = {}
    tokens: list[str] = []
    buffer = b""

    def consume_line(line_bytes: bytes) -> None:
        nonlocal meta
        line = line_bytes.decode("utf-8", errors="replace").strip()
        if not line or line.startswith("event:") or not line.startswith("data:"):
            return

        payload = line[5:].strip()
        if not payload or payload == "[DONE]":
            return
        try:
            data = json.loads(payload)
        except json.JSONDecodeError:
            return

        if data.get("type") == "chat_meta":
            meta = data
            print(
                f"\n[检索耗时 {meta.get('retrieval_ms', 0)}ms]  "
                f"检索词: {meta.get('retrieval_query', '')}"
            )
            for source in meta.get("sources", []):
                print(
                    f"  [{source.get('number', '')}] "
                    f"{source.get('title', '')}"
                )
            print()
            return

        choices = data.get("choices")
        if not isinstance(choices, list) or not choices:
            return
        delta = choices[0].get("delta", {})
        token = delta.get("content", "") if isinstance(delta, dict) else ""
        if token:
            tokens.append(str(token))
            print(token, end="", flush=True)

    try:
        for raw_chunk in response.iter_content(chunk_size=None):
            if not raw_chunk:
                continue
            buffer += raw_chunk
            while b"\n" in buffer:
                line_bytes, buffer = buffer.split(b"\n", 1)
                consume_line(line_bytes)
        if buffer.strip():
            consume_line(buffer)
    except requests.RequestException as exc:
        print()
        return {
            "answer": "".join(tokens),
            "sources": meta.get("sources", []),
            "retrieval_query": meta.get("retrieval_query", ""),
            "retrieval_ms": meta.get("retrieval_ms", 0),
            "error": str(exc),
        }
    finally:
        response.close()

    print()
    return {
        "answer": "".join(tokens),
        "sources": meta.get("sources", []),
        "retrieval_query": meta.get("retrieval_query", ""),
        "retrieval_ms": meta.get("retrieval_ms", 0),
        "error": "",
    }


def main() -> None:
    args = parse_args()
    global BASE_URL, PROJECT_PATH
    BASE_URL = args.base_url.rstrip("/")
    PROJECT_PATH = args.project_path

    input_rows = load_input_rows(Path(args.input), args.question_column)
    selected_rows = input_rows[args.start - 1 :]
    if args.limit is not None:
        selected_rows = selected_rows[: args.limit]
    if not selected_rows:
        raise SystemExit("没有找到待评估的问题。")

    timestamp = time.strftime("%Y%m%d_%H%M%S")
    output_path = add_timestamp(Path(args.output), timestamp)
    if args.jsonl_output:
        jsonl_path = add_timestamp(Path(args.jsonl_output), timestamp)
    else:
        jsonl_path = output_path.with_suffix(".jsonl")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    jsonl_path.parent.mkdir(parents=True, exist_ok=True)

    results: list[dict[str, Any]] = []
    total = len(selected_rows)
    batch_started = time.perf_counter()

    with jsonl_path.open("w", encoding="utf-8") as jsonl_file:
        for position, source_row in enumerate(selected_rows, start=1):
            input_index = args.start + position - 1
            question = source_row["question"]
            print(f"\n{'=' * 80}")
            print(f"[{position}/{total}] 问：{question}")

            started = time.perf_counter()
            response = chat_stream(question, top_k=args.top_k)
            elapsed_ms = round((time.perf_counter() - started) * 1000, 2)
            row = build_result_row(input_index, source_row, response, elapsed_ms)
            results.append(row)
            jsonl_file.write(json.dumps(row, ensure_ascii=False) + "\n")
            jsonl_file.flush()

            status = "OK" if row["ok"] else "ERROR"
            print(
                f"[{position}/{total}] {status} 端到端耗时={elapsed_ms:.0f}ms "
                f"检索耗时={row['retrieval_ms']}ms 来源={row['source_count']}"
            )
            if args.sleep > 0 and position < total:
                time.sleep(args.sleep)

    write_excel(output_path, results)
    total_ms = round((time.perf_counter() - batch_started) * 1000, 2)
    success_count = sum(1 for row in results if row["ok"])
    print(f"\n完成：{success_count}/{len(results)} 条成功，总耗时 {total_ms:.0f}ms")
    print(f"Excel：{output_path}")
    print(f"JSONL：{jsonl_path}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="通过 LLM Wiki chat_stream() 批量生成评估答案并导出 Excel/JSONL。"
    )
    parser.add_argument("--input", default="/Users/mickey/project/PA-ALG/llm-wiki-rag/docs/评估集.xlsx", help="输入文件：XLSX、CSV、JSON、JSONL 或 TXT。")
    parser.add_argument(
        "--output",
        default="/Users/mickey/project/PA-ALG/llm-wiki-rag/docs/评估集_llm_wiki_chat_stream_result.xlsx",
        help="结果 Excel 路径，实际文件名会自动追加时间戳。",
    )
    parser.add_argument(
        "--jsonl-output",
        help="结果 JSONL 路径，实际文件名会自动追加时间戳；默认与 Excel 同名。",
    )
    parser.add_argument("--question-column", default="question", help="问题列名。")
    parser.add_argument("--top-k", type=int, default=8, help="每题检索的 chunk 数量。")
    parser.add_argument("--start", type=int, default=1, help="从第几条数据开始，1 表示第一条。")
    parser.add_argument("--limit", type=int, help="最多评估多少条。")
    parser.add_argument("--sleep", type=float, default=0.2, help="题目之间的等待秒数。")
    parser.add_argument("--base-url", default=BASE_URL, help="LLM Wiki 服务地址。")
    parser.add_argument("--project-path", default=PROJECT_PATH, help="LLM Wiki 项目路径。")
    args = parser.parse_args()
    if args.start < 1:
        parser.error("--start 必须大于等于 1")
    if args.limit is not None and args.limit < 1:
        parser.error("--limit 必须大于等于 1")
    if args.top_k < 1:
        parser.error("--top-k 必须大于等于 1")
    return args


def add_timestamp(path: Path, timestamp: str) -> Path:
    suffix = path.suffix
    if suffix:
        return path.with_name(f"{path.stem}_{timestamp}{suffix}")
    return path.with_name(f"{path.name}_{timestamp}")


def load_input_rows(path: Path, question_column: str) -> list[dict[str, str]]:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        raw_rows = load_xlsx_rows(path)
    elif suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as file:
            raw_rows = [dict(row) for row in csv.DictReader(file)]
    elif suffix == ".jsonl":
        raw_rows = [
            json.loads(line)
            for line in path.read_text(encoding="utf-8-sig").splitlines()
            if line.strip()
        ]
    elif suffix == ".json":
        payload = json.loads(path.read_text(encoding="utf-8-sig"))
        raw_rows = payload.get("data", []) if isinstance(payload, dict) else payload
    elif suffix in {"", ".txt"}:
        raw_rows = [
            {question_column: line.strip()}
            for line in path.read_text(encoding="utf-8-sig").splitlines()
            if line.strip()
        ]
    else:
        raise ValueError(f"不支持的输入格式：{path.suffix}")

    if not isinstance(raw_rows, list) or any(not isinstance(row, dict) for row in raw_rows):
        raise ValueError("JSON 输入必须是对象数组，或包含 data 对象数组。")
    if not raw_rows:
        return []

    columns = list(raw_rows[0].keys())
    question_key = resolve_column(columns, question_column, QUESTION_ALIASES, required=True)
    gold_key = resolve_column(columns, "", GOLD_ANSWER_ALIASES)
    category_key = resolve_column(columns, "", CATEGORY_ALIASES)
    actual_key = resolve_column(
        columns,
        "",
        ("actual_answer", "answer", "model_answer", "模型回答", "实际回答", "回答"),
    )

    rows: list[dict[str, str]] = []
    for raw_row in raw_rows:
        question = stringify(raw_row.get(question_key)).strip()
        if not question:
            continue
        rows.append(
            {
                "question": question,
                "gold_answer": stringify(raw_row.get(gold_key)).strip() if gold_key else "",
                "category": stringify(raw_row.get(category_key)).strip() if category_key else "",
                "previous_actual_answer": (
                    stringify(raw_row.get(actual_key)).strip() if actual_key else ""
                ),
            }
        )
    return rows


def load_xlsx_rows(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    values = sheet.iter_rows(values_only=True)
    headers = [stringify(value).strip() for value in next(values, [])]
    rows: list[dict[str, Any]] = []
    for values_row in values:
        row = {
            header: values_row[index] if index < len(values_row) else None
            for index, header in enumerate(headers)
            if header
        }
        if any(stringify(value).strip() for value in row.values()):
            rows.append(row)
    workbook.close()
    return rows


def resolve_column(
    columns: list[str],
    preferred: str,
    aliases: tuple[str, ...],
    required: bool = False,
) -> str | None:
    normalized = {str(column).strip().lower(): str(column) for column in columns}
    for candidate in (preferred, *aliases):
        if not candidate:
            continue
        found = normalized.get(candidate.strip().lower())
        if found:
            return found
    if required:
        raise ValueError(f"找不到问题列 {preferred!r}，可用列：{columns}")
    return None


def build_result_row(
    input_index: int,
    source_row: dict[str, str],
    response: dict[str, Any],
    elapsed_ms: float,
) -> dict[str, Any]:
    sources = response.get("sources", [])
    if not isinstance(sources, list):
        sources = []
    error = stringify(response.get("error")).strip()
    contexts = [format_source_context(source) for source in sources if isinstance(source, dict)]
    return {
        "input_index": input_index,
        "question": source_row["question"],
        "actual_answer": stringify(response.get("answer")),
        "gold_answer": source_row.get("gold_answer", ""),
        "contexts": contexts,
        "category": source_row.get("category", ""),
        "previous_actual_answer": source_row.get("previous_actual_answer", ""),
        "retrieval_query": stringify(response.get("retrieval_query")),
        "retrieval_ms": number_or_zero(response.get("retrieval_ms")),
        "elapsed_ms": elapsed_ms,
        "source_count": len(sources),
        "sources": sources,
        "ok": not error,
        "error": error,
    }


def format_source_context(source: dict[str, Any]) -> str:
    number = source.get("number", "")
    title = stringify(source.get("title")).strip()
    path = stringify(source.get("path")).strip()
    score = source.get("score")
    parts = [f"[{number}] {title}".strip()]
    if path:
        parts.append(f"path={path}")
    if score is not None:
        parts.append(f"score={score}")
    return " | ".join(parts)


def write_excel(path: Path, rows: list[dict[str, Any]]) -> None:
    fields = [
        "input_index",
        "question",
        "actual_answer",
        "gold_answer",
        "contexts",
        "category",
        "previous_actual_answer",
        "retrieval_query",
        "retrieval_ms",
        "elapsed_ms",
        "source_count",
        "sources",
        "ok",
        "error",
    ]
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "评估结果"

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for column_index, field in enumerate(fields, start=1):
        cell = sheet.cell(1, column_index, field)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_index, row in enumerate(rows, start=2):
        for column_index, field in enumerate(fields, start=1):
            value = row.get(field, "")
            if isinstance(value, (list, dict)):
                value = json.dumps(value, ensure_ascii=False)
            cell = sheet.cell(row_index, column_index, value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = {
        "input_index": 12,
        "question": 36,
        "actual_answer": 70,
        "gold_answer": 45,
        "contexts": 55,
        "category": 16,
        "previous_actual_answer": 55,
        "retrieval_query": 40,
        "retrieval_ms": 14,
        "elapsed_ms": 14,
        "source_count": 13,
        "sources": 65,
        "ok": 10,
        "error": 40,
    }
    for column_index, field in enumerate(fields, start=1):
        sheet.column_dimensions[get_column_letter(column_index)].width = widths[field]
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    summary = workbook.create_sheet("汇总")
    success_count = sum(1 for row in rows if row["ok"])
    retrieval_values = [row["retrieval_ms"] for row in rows if row["ok"]]
    elapsed_values = [row["elapsed_ms"] for row in rows if row["ok"]]
    summary_rows = [
        ("指标", "值"),
        ("总题数", len(rows)),
        ("成功数", success_count),
        ("失败数", len(rows) - success_count),
        ("成功率", success_count / len(rows) if rows else 0),
        ("平均检索耗时(ms)", average(retrieval_values)),
        ("平均端到端耗时(ms)", average(elapsed_values)),
        ("平均来源数", average([row["source_count"] for row in rows if row["ok"]])),
    ]
    for row_index, values in enumerate(summary_rows, start=1):
        for column_index, value in enumerate(values, start=1):
            summary.cell(row_index, column_index, value)
    for cell in summary[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    summary["B5"].number_format = "0.0%"
    summary.column_dimensions["A"].width = 24
    summary.column_dimensions["B"].width = 18

    workbook.save(path)


def average(values: list[float | int]) -> float:
    return round(sum(values) / len(values), 2) if values else 0


def number_or_zero(value: Any) -> float | int:
    return value if isinstance(value, (float, int)) and not isinstance(value, bool) else 0


def stringify(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


if __name__ == "__main__":
    main()
